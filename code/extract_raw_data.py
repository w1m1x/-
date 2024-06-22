import pandas as pd
import openpyxl
import os


def append_excel(data, excelname, sheetname, insert_type):
    """
    将DataFrame数据追加到Excel文件中指定的表。
    """
    if not os.path.exists(excelname):
        print(f"文件 {excelname} 不存在，创建新文件。")
        # 如果文件不存在，创建一个新的工作簿
        wb = openpyxl.Workbook()
        wb.save(excelname)

    try:
        # 加载现有工作簿
        book = openpyxl.load_workbook(excelname)

        if sheetname not in book.sheetnames:
            print(f"工作表 {sheetname} 不存在，创建新的工作表。")
            # 如果指定的工作表不存在，创建一个新的工作表
            book.create_sheet(title=sheetname)

        sheet = book[sheetname]

        # 获取原数据的行数
        original_row = sheet.max_row

        if insert_type == 'w':  # 选择写入excel数据方式，w为覆盖模式，a+为追加模式
            startrow = 0  # 覆盖模式从头开始写入
        elif insert_type == 'a+':
            startrow = original_row  # 追加模式从原数据行数之后开始写入

        # 如果是覆盖模式，需要清空原有数据
        if insert_type == 'w':
            for row in sheet.iter_rows():
                for cell in row:
                    cell.value = None

        # 将DataFrame数据写入到指定的起始行
        for row_idx, row_data in enumerate(data.values, start=startrow + 1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        # 保存工作簿
        book.save(excelname)
        book.close()
        print(f"数据成功写入 {excelname} 的 {sheetname} 表中。")
    except Exception as e:
        print(f"Error writing to {excelname}: {e}")


if __name__ == '__main__':
    try:
        # 检查并处理HTML文件
        html_file_path = "./raw_relation.html"
        if not os.path.exists(html_file_path):
            print(f"HTML 文件 {html_file_path} 不存在。")
        else:
            # 读取HTML表格
            df_list = pd.read_html(html_file_path, encoding='utf-8', header=0)
            num_tables = len(df_list)
            print(f"在HTML文件中找到 {num_tables} 个表格。")

            # 追加数据到Excel
            excel_file_path = './raw_data.xlsx'
            for i in range(num_tables):
                append_excel(df_list[i], excel_file_path, 'Sheet1', 'a+')
    except Exception as e:
        print(f"Error processing HTML tables: {e}")
